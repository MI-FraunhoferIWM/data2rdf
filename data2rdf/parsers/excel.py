"""Data2rdf excel parser"""

import warnings
from io import BytesIO
from typing import Any, Dict, List, Union
from urllib.parse import quote, urljoin

import pandas as pd
from openpyxl import load_workbook
from pydantic import Field

from data2rdf.models.graph import PropertyGraph, QuantityGraph
from data2rdf.utils import make_prefix
from data2rdf.warnings import MappingMissmatchWarning

from .base import ABoxBaseParser, BaseFileParser, TBoxBaseParser

from .utils import (  # isort:skip
    _make_tbox_classes,
    _make_tbox_json_ld,
    _strip_unit,
    _value_exists,
)

from data2rdf.models.mapping import (  # isort:skip
    ABoxExcelMapping,
    TBoxBaseMapping,
    CustomRelationPropertySubgraph,
    CustomRelationQuantitySubgraph,
)


def _load_data_file(
    self: "Union[ExcelTBoxParser, ExcelABoxParser]",
) -> BytesIO:
    """Load excel file as bytes"""
    if isinstance(self.raw_data, str):
        with open(self.raw_data, mode="rb") as file:
            content = BytesIO(file.read())
    elif isinstance(self.raw_data, bytes):
        content = BytesIO(self.raw_data)
    else:
        raise TypeError(
            f"`raw_data` must be of type `str` or `btyes`, not `{type(self.raw_data)}`"
        )
    return content


class ExcelTBoxParser(TBoxBaseParser):
    """
    Parses a data file of type excel in b box mode
    """

    sheet: str = Field(..., description="Name of the sheet for the mapping.")

    header_length: int = Field(
        1, description="Length of the header of the excel sheet", ge=1
    )

    # OVERRIDE
    mapping: Union[str, List[TBoxBaseMapping]] = Field(
        ...,
        description="""File path to the mapping file to be parsed or
        a list with the mapping.""",
    )

    # OVERRIDE
    @property
    def mapping_model(self) -> TBoxBaseMapping:
        "TBox Mapping Model"
        return TBoxBaseMapping

    # OVERRIDE
    @property
    def json_ld(self) -> "Dict[str, Any]":
        """Make the json-ld if pipeline is in abox-mode"""
        return _make_tbox_json_ld(self)

    # OVERRIDE
    @classmethod
    def _run_parser(
        cls,
        self: "ExcelTBoxParser",
        datafile: BytesIO,
        mapping: "List[TBoxBaseMapping]",
    ) -> None:
        """
        Run excel parser in tbox mode.

        Parameters
        ----------
        self : ExcelTBoxParser
            The instance of the parser.
        datafile : BytesIO
            The excel file to be parsed.
        mapping : List[TBoxBaseMapping]
            The list of mappings to be applied.

        Returns
        -------
        None
            This function does not return any value.
        """
        df = pd.read_excel(datafile, sheet_name=self.sheet)
        _make_tbox_classes(self, df, mapping)

    # OVERRIDE
    @classmethod
    def _load_data_file(cls, self: "ExcelTBoxParser") -> BytesIO:
        """Load excel file"""
        return _load_data_file(self)


class ExcelABoxParser(ABoxBaseParser):
    """
    Parses a data file of type excel in a box mode
    """

    unit_from_macro: bool = Field(
        False,
        description="When enabled, units coming from excel macros will be taken into account.",
    )

    unit_macro_location: int = Field(
        -1,
        description="Index where the marco for the unit in an excel cell might be located.",
    )

    # OVERRIDE
    mapping: Union[str, List[ABoxExcelMapping]] = Field(
        ...,
        description="""File path to the mapping file to be parsed or
        a list with the mapping.""",
    )

    # OVERRIDE
    @property
    def mapping_model(self) -> ABoxExcelMapping:
        "Mapping Model"
        return ABoxExcelMapping

    # OVERRIDE
    @property
    def json_ld(self) -> Dict[str, Any]:
        """
        Returns the JSON-LD representation of the data in ABox mode.

        The JSON-LD is constructed based on the metadata and dataframe data.
        If the file description is not suppressed, it includes the metadata and dataframe data tables.
        Otherwise, it returns a list of JSON-LD representations of the individual models.

        :return: A dictionary representing the JSON-LD data.
        """

        if not self.config.suppress_file_description:
            tables = []

            if self.general_metadata:
                meta_table = {
                    "@type": "csvw:Table",
                    "rdfs:label": "Metadata",
                    "csvw:row": [],
                }

                for mapping in self.general_metadata:
                    if isinstance(mapping, QuantityGraph):
                        row = {
                            "@type": "csvw:Row",
                            "csvw:titles": {
                                "@type": "xsd:string",
                                "@value": mapping.key,
                            },
                            "qudt:quantity": mapping.json_ld,
                        }
                        meta_table["csvw:row"].append(row)
                    elif isinstance(mapping, PropertyGraph):
                        row = {
                            "@type": "csvw:Row",
                            "csvw:titles": {
                                "@type": "xsd:string",
                                "@value": mapping.key,
                            },
                            "csvw:describes": mapping.json_ld,
                        }
                        meta_table["csvw:row"].append(row)
                    else:
                        raise TypeError(
                            f"Mapping must be of type {QuantityGraph} or {PropertyGraph}, not {type(mapping)}"
                        )
                tables += [meta_table]

            if self.dataframe_metadata:
                column_schema = {"@type": "csvw:Schema", "csvw:column": []}
                tables += [
                    {
                        "@type": "csvw:Table",
                        "rdfs:label": "Dataframe",
                        "csvw:tableSchema": column_schema,
                    }
                ]
                for idx, mapping in enumerate(self.dataframe_metadata):
                    if isinstance(mapping, QuantityGraph):
                        entity = {"qudt:quantity": mapping.json_ld}
                    elif isinstance(mapping, PropertyGraph):
                        entity = {"dcterms:subject": mapping.json_ld}
                    else:
                        raise TypeError(
                            f"Mapping must be of type {QuantityGraph} or {PropertyGraph}, not {type(mapping)}"
                        )

                    if self.config.data_download_uri:
                        download_url = {
                            "dcterms:identifier": {
                                "@type": "xsd:anyURI",
                                "@value": urljoin(
                                    str(self.config.data_download_uri),
                                    f"column-{idx}",
                                ),
                            }
                        }
                    else:
                        download_url = {}

                    column = {
                        "@type": "csvw:Column",
                        "csvw:titles": {
                            "@type": "xsd:string",
                            "@value": mapping.key,
                        },
                        **entity,
                        "foaf:page": {
                            "@type": "foaf:Document",
                            "dcterms:format": {
                                "@type": "xsd:anyURI",
                                "@value": "https://www.iana.org/assignments/media-types/application/json",
                            },
                            "dcterms:type": {
                                "@type": "xsd:anyURI",
                                "@value": "http://purl.org/dc/terms/Dataset",
                            },
                            **download_url,
                        },
                    }
                    column_schema["csvw:column"].append(column)

            # flatten list if only one value exists
            if len(tables) == 1:
                tables = tables.pop()
            # make relation to csvw:table property
            if tables:
                csvw_tables = {"csvw:table": tables}
            else:
                csvw_tables = {}

            json_ld = {
                "@context": {
                    f"{self.config.prefix_name}": make_prefix(self.config),
                    "csvw": "http://www.w3.org/ns/csvw#",
                    "rdfs": "http://www.w3.org/2000/01/rdf-schema#",
                    "dcat": "http://www.w3.org/ns/dcat#",
                    "xsd": "http://www.w3.org/2001/XMLSchema#",
                    "dcterms": "http://purl.org/dc/terms/",
                    "qudt": "http://qudt.org/schema/qudt/",
                    "csvw": "http://www.w3.org/ns/csvw#",
                    "foaf": "http://xmlns.com/foaf/spec/",
                },
                "@id": f"{self.config.prefix_name}:tableGroup",
                "@type": "csvw:TableGroup",
                **csvw_tables,
            }
        else:
            json_ld = {
                "@graph": [model.json_ld for model in self.general_metadata]
                + [model.json_ld for model in self.dataframe_metadata]
            }
        return json_ld

    # OVERRIDE
    @classmethod
    def _run_parser(
        cls,
        self: "ExcelABoxParser",
        datafile: BytesIO,
        mapping: "List[ABoxExcelMapping]",
    ) -> None:
        """
        Parses the metadata, dataframe metadata, and dataframe from an Excel file.

        Args:
            self (ExcelABoxParser): The instance of the ExcelABoxParser class.
            datafile (BytesIO): The file object containing the Excel file.
            mapping (List[ABoxExcelMapping]): The list of mappings to use for parsing.

        Returns:
            None: This function does not return anything.
        """

        workbook = load_workbook(filename=datafile, data_only=True)
        datafile.seek(0)
        macros = load_workbook(filename=datafile)
        datafile.seek(0)

        self._general_metadata = []
        self._dataframe_metadata = []
        self._dataframe = {}

        for datum in mapping:
            worksheet = workbook[datum.worksheet]

            if datum.suffix_from_location:
                suffix = worksheet[datum.suffix].value
                if not suffix:
                    suffix = datum.suffix
                    message = f"""Could not properly resolve suffix location `{datum.suffix}`
                                  Will use the location itself as suffix.
                                """
                    warnings.warn(message, MappingMissmatchWarning)
            else:
                suffix = datum.suffix
            suffix = quote(suffix)

            if not datum.custom_relations:
                if datum.value_location and datum.dataframe_start:
                    raise RuntimeError(
                        """Both, `value_location` and `dataframe_start
                        are set. Only one of them must be set."""
                    )

                # find data for dataframe
                if datum.dataframe_start:
                    column_name = datum.dataframe_start.rstrip("0123456789")
                    dataframe_end = f"{column_name}{worksheet.max_row}"

                    column = worksheet[datum.dataframe_start : dataframe_end]
                    if column:
                        self._dataframe[suffix] = [
                            cell[0].value for cell in column
                        ]
                    else:
                        message = f"""Concept with key `{datum.key}`
                                    does not have a dataframe from `{datum.dataframe_start}`
                                    until `{dataframe_end}` .
                                    Concept will be omitted in graph.
                                    """
                        warnings.warn(message, MappingMissmatchWarning)

                # check if there is a macro for the unit of the entity
                if self.unit_from_macro and datum.value_location:
                    macro_worksheet = macros[datum.worksheet]
                    macro_value_cell = macro_worksheet[
                        datum.value_location
                    ].number_format.split()
                    if len(macro_value_cell) != 1:
                        macro_unit = macro_value_cell[self.unit_macro_location]
                    else:
                        macro_unit = None
                else:
                    macro_unit = None

                # check if there is a unit somewhere in the sheet
                if datum.unit_location:
                    unit_location = worksheet[datum.unit_location].value
                    if not unit_location:
                        message = f"""Concept with key `{datum.key}`
                                    does not have a unit at location `{datum.unit_location}`.
                                    This mapping for the unit will be omitted in graph.
                                    """
                        warnings.warn(message, MappingMissmatchWarning)
                else:
                    unit_location = None

                # decide which unit to take
                unit = datum.unit or unit_location or macro_unit
                if unit:
                    if not isinstance(unit, str):
                        raise TypeError(
                            f"""Unit `{unit}` for key `{datum.key}` is not a string.
                            Is it a bad mapping?"""
                        )
                    unit = _strip_unit(unit, self.config.remove_from_unit)

                # make model
                model_data = {
                    "key": datum.key,
                    "unit": unit,
                    "iri": datum.iri,
                    "suffix": suffix,
                    "annotation": datum.annotation,
                    "config": self.config,
                }

                if datum.value_location and not datum.dataframe_start:
                    value = worksheet[datum.value_location].value

                    if model_data.get("unit") and _value_exists(value):
                        model_data["value"] = value
                    elif not model_data.get("unit") and _value_exists(value):
                        model_data["value"] = str(value)
                    else:
                        message = f"""Concept with key `{datum.key}`
                                    does not have a value at location `{datum.value_location}`.
                                    Concept will be omitted in graph.
                                    """
                        warnings.warn(message, MappingMissmatchWarning)
                else:
                    value = None

                value_exists = _value_exists(value)

                if value_exists or suffix in self.dataframe:
                    if datum.value_relation:
                        model_data["value_relation"] = datum.value_relation
                    if model_data.get("unit"):
                        if datum.unit_relation:
                            model_data["unit_relation"] = datum.unit_relation
                        model = QuantityGraph(**model_data)
                    else:
                        model = PropertyGraph(
                            **model_data,
                            value_datatype=datum.value_datatype,
                            value_relation_type=datum.value_relation_type,
                        )

                    if value_exists:
                        self._general_metadata.append(model)
                    else:
                        self._dataframe_metadata.append(model)

            else:
                for relation in datum.custom_relations:
                    value = worksheet[relation.object_location].value

                    if isinstance(
                        relation.object_data_type,
                        (
                            CustomRelationPropertySubgraph,
                            CustomRelationQuantitySubgraph,
                        ),
                    ):
                        if isinstance(
                            relation.object_data_type,
                            CustomRelationPropertySubgraph,
                        ):
                            Model = PropertyGraph
                        else:
                            Model = QuantityGraph
                        model = Model(
                            value=value,
                            **relation.object_data_type.model_dump(),
                        )
                        model.suffix += "_" + suffix
                        model = PropertyGraph(
                            value_relation=relation.relation,
                            value_relation_type="object_property",
                            value=model,
                            iri=datum.iri,
                            suffix=suffix,
                            config=self.config,
                        )
                        self._general_metadata.append(model)
                    elif _value_exists(value):
                        model = PropertyGraph(
                            value_relation=relation.relation,
                            value_relation_type=relation.relation_type,
                            value_datatype=relation.object_data_type,
                            value=value,
                            iri=datum.iri,
                            suffix=suffix,
                            config=self.config,
                        )
                        self._general_metadata.append(model)
                    else:
                        message = f"""Concept with for iri `{datum.iri}`
                                        does not have a value at location `{relation.object_location}`.
                                        Concept will be omitted in graph.
                                        """
                        warnings.warn(message, MappingMissmatchWarning)

        # set dataframe as pd dataframe
        self._dataframe = pd.DataFrame.from_dict(
            self._dataframe, orient="index"
        ).transpose()
        # check if drop na:
        if self.dropna:
            self._dataframe.dropna(how="all", inplace=True)

    # OVERRIDE
    @classmethod
    def _load_data_file(cls, self: "ExcelABoxParser") -> BytesIO:
        """Load excel file"""
        return _load_data_file(self)


class ExcelParser(BaseFileParser):
    """Parser for excel files"""

    # OVERRIDE
    @property
    def _abox_parser(self) -> ExcelABoxParser:
        """Pydantic Model for Excel ABox parser"""
        return ExcelABoxParser

    # OVERRIDE
    @property
    def _tbox_parser(self) -> ExcelTBoxParser:
        """Pydantic Model for Excel TBox parser"""
        return ExcelTBoxParser

    # OVERRIDE
    @property
    def media_type(self) -> str:
        """IANA Media type definition of the resource to be parsed."""
        return "https://www.iana.org/assignments/media-types/application/vnd.ms-excel"
