import pandas as pd
from openpyxl import load_workbook

from data2rdf.parser import DataParser


class ExcelParser(DataParser):
    """
    Generates the excel input sheet that can be used
    as input for abox skeleton files.

    Attributes:
        f_path (str): The file path for the csv file used as input for the parser. This path gets also stored as dcat:downloadURL attribute for the created dcat:Dataset individual by the rdf_generation class.
        location_mapping_f_path (str): Path to the excel file, that holds the location of the meta data and column data cells that should be extracted.
        server_f_path (str): By default the file path for the csv file (f_path) gets used as dcat:downloadURL attribute for the created dcat:Dataset individual. On a server the actual download url of the file should be used
        (e.g. on the DSMS https://127.0.0.1/api/knowledge/data-files/764f6e51-a244-42f9-a754-c3e2861f63e4/raw_data/excel_file.xlsx).
        data_storage_path (str): Optional different storage location for the hdf5 file holding the data. Default is the same location as the input file.
        data_storage_group_name (str): Name of the group in the hdf5 to store the data. Using the data_storage_path and the data_storage_group_name multiple datasets can be stored in the same hdf5 file.
        namespace (str): The namespace that will be used by the rdf_generation class to construct the abox individuals.

    """

    def __init__(
        self,
        f_path,
        location_mapping_f_path,
        # when a file is stored on a server it makes sense to store this path
        # (e.g. http access) instead of the local path
        server_f_path=None,
        # the columns get stored in a dedicated format (hdf5 by default)
        data_storage_path=None,
        # one hdf5 store is probaby enough for our current setup (use different
        # groups (folders), for each file)
        data_storage_group_name="df",
        namespace="http://www.test.de",
    ):
        super().__init__(
            f_path,
            server_f_path,
            data_storage_path,
            data_storage_group_name,
            namespace,
        )

        self.location_mapping_f_path = location_mapping_f_path

    def parser_data(self):
        self.load_file()
        self.generate_file_uuid()
        self.load_mapping_file()
        self.parse_meta_data()
        self.generate_column_df()
        self.parse_table()
        self.generate_file_meta_df()

    def load_mapping_file(self):
        self.meta_mapping_df = pd.read_excel(
            self.location_mapping_f_path,
            engine="openpyxl",
            sheet_name="meta",
        )
        self.meta_mapping_df.fillna("", inplace=True)
        self.column_mapping_df = pd.read_excel(
            self.location_mapping_f_path,
            engine="openpyxl",
            sheet_name="columns",
        )
        self.column_mapping_df.fillna("", inplace=True)

    def load_file(self):
        # self.file = open(self.f_path, 'r', encoding=self.encoding).read()
        self.workbook = load_workbook(filename=self.f_path, data_only=True)
        # some datasets do have the unit encoded as macro of the cell, in order
        # to extract this unit, the unformatted workbook is needed
        self.workbook_macros = load_workbook(filename=self.f_path)

    def parse_meta_data(self):
        """
        Extracts meta data from the excel worksheet using the location mapping
        information from the meta_mapping_df
        """

        meta_data = []
        for _, row in self.meta_mapping_df.iterrows():
            if not row["Excel worksheet"]:
                continue

            # get worksheet from mapping
            worksheet = self.workbook[row["Excel worksheet"]]

            cell_location_name = row["Variable name cell location"]
            cell_location_value = row["Variable value cell location"]
            cell_location_unit = row["Variable unit cell location"]

            if cell_location_value == "":
                continue

            meta_name = worksheet[cell_location_name].value
            meta_value = worksheet[cell_location_value].value

            if cell_location_unit == "":
                macro_worksheet = self.workbook_macros[row["Excel worksheet"]]
                macro_value_cell = macro_worksheet[
                    cell_location_value
                ].number_format

                if len(macro_value_cell.split(" ")) != 1:
                    meta_unit = macro_value_cell.split(" ")[1].strip('"')
                else:
                    meta_unit = ""

            else:
                meta_unit = worksheet[cell_location_unit].value

            meta_name = meta_name.strip(":").strip("=")

            meta_data.append(
                {"index": meta_name, "value": meta_value, "unit": meta_unit}
            )

        self.meta_df = pd.DataFrame(meta_data)

    def split_meta_df(self):
        """
        Split the meta data into basic description and quantities
        """
        self.meta_quantity_df = self.meta_df.loc[
            (self.meta_df["unit"] != ""), :
        ]
        self.meta_description_df = self.meta_df.loc[
            (self.meta_df["unit"] == ""), :
        ]

    def parse_table(self):
        self.df_table = pd.DataFrame()
        for _, row in self.column_df.iterrows():
            # print(self.workbook[row["worksheet"]][row["start"]:row["end"]])
            column = [
                cell[0].value
                for cell in self.workbook[row["worksheet"]][
                    row["start"] : row["end"]
                ]
            ]
            if not len(column):
                column = ""
            self.df_table[row["index"]] = column

        self.df_table = self.df_table.apply(pd.to_numeric, errors="coerce")

        # print(self.df_table)

    def generate_column_df(self):
        """
        Extracts meta data from the excel worksheet using the location mapping
        information from the meta_mapping_df
        """

        column_data = []
        for _, row in self.column_mapping_df.iterrows():
            sheet = row["Excel worksheet"]

            if not sheet:
                continue

            # get worksheet from mapping
            worksheet = self.workbook[sheet]

            # get name and unit from location
            cell_location_name = row["Column name cell location"]
            cell_location_unit = row["Column unit cell location"]

            col_name = worksheet[cell_location_name].value

            if cell_location_unit == "":
                col_unit = ""
            else:
                col_unit = worksheet[cell_location_unit].value

            col_unit = col_unit.strip("[").strip("]")
            col_name = col_name.strip(":").strip("=")

            # auto fill column end
            # column names are always letters, row names are always numbers
            column_name = row["Column cell location start"].rstrip(
                "0123456789"
            )
            row_name = int(
                row["Column cell location start"].lstrip(column_name)
            )

            # go trough the rows and check when the value is empty to get the
            # location end
            while True:
                cell_name = f"{column_name}{row_name}"
                cell = worksheet[cell_name]

                if cell.value == "" or cell.value is None:
                    row_name -= 1
                    col_end_pos = f"{column_name}{row_name}"
                    break

                row_name += 1

            # get column start from location
            col_start_pos = row["Column cell location start"]

            column_data.append(
                {
                    "index": col_name,
                    "titles": col_name,
                    "unit": col_unit,
                    "start": col_start_pos,
                    "end": col_end_pos,
                    "worksheet": sheet,
                }
            )

        self.column_df = pd.DataFrame(column_data)
        # print(self.column_df)

    # def clean_table_df(self):
    #    self.df_table = self.df_table.iloc[1:,:]

    def generate_excel_spreadsheet(self, output_path):
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            self.file_meta_df.to_excel(writer, "file")
            self.column_df.to_excel(writer, "column_meta")
            self.meta_df.to_excel(writer, "meta")

    def generate_data_storage(self):
        # print(self.df_table)
        self.df_table.to_hdf(
            self.data_storage_path, key=self.data_storage_group_name
        )
